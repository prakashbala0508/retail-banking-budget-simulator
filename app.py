
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import requests
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(
    page_title="Retail Banking Budget Simulator",
    page_icon="Bank",
    layout="wide"
)

@st.cache_data(ttl=3600)
def pull_fred(series_id, col_name):
    url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
    try:
        df = pd.read_csv(url)
        df.columns = ["date", col_name]
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        return df.dropna()
    except:
        return None

AVG_LOANS_M        = 8_500.0
AVG_DEPOSITS_M     = 11_200.0
NIM_BASE           = 0.031
NONINT_INCOME_M    = 420.0
EFFICIENCY_RATIO   = 0.61
CHARGE_OFF_RATE    = 0.0045
TAX_RATE           = 0.25
FTE_PER_BILLION    = 21
AVG_COMP_PER_FTE_K = 85.0

net_interest_income_M = AVG_LOANS_M * NIM_BASE
total_revenue_M       = net_interest_income_M + NONINT_INCOME_M
nonint_expense_M      = total_revenue_M * EFFICIENCY_RATIO
provision_M           = AVG_LOANS_M * CHARGE_OFF_RATE
pretax_M              = total_revenue_M - nonint_expense_M - provision_M
net_income_base_M     = pretax_M * (1 - TAX_RATE)
fte_base              = round((AVG_LOANS_M / 1_000) * FTE_PER_BILLION)
total_comp_base_M     = (fte_base * AVG_COMP_PER_FTE_K) / 1_000

def project_scenario(nim_change_bps, loan_growth_rate, gdp_growth,
                     unemployment_delta, base_charge_off, n_years=3):
    loans      = AVG_LOANS_M
    nim        = NIM_BASE
    charge_off = base_charge_off
    eff_ratio  = EFFICIENCY_RATIO
    nonint_inc = NONINT_INCOME_M
    loans_prev = AVG_LOANS_M
    rows = []
    for yr in range(1, n_years + 1):
        loans      = loans * (1 + loan_growth_rate)
        nim        = max(nim + nim_change_bps / 10_000, 0.020)
        nii        = loans * nim
        nonint_inc = nonint_inc * (1 + gdp_growth)
        charge_off = max(charge_off + unemployment_delta * 0.001, 0.002)
        provision  = loans * charge_off
        total_rev  = nii + nonint_inc
        eff_ratio  = max(eff_ratio - 0.005, 0.50)
        nonint_exp = total_rev * eff_ratio
        pretax     = total_rev - nonint_exp - provision
        net_inc    = pretax * (1 - TAX_RATE)
        fte        = round((loans_prev / 1_000) * FTE_PER_BILLION)
        loans_prev = loans
        total_comp = (fte * AVG_COMP_PER_FTE_K) / 1_000
        rows.append({
            "Year":                             f"Year {yr}",
            "Avg Loans ($M)":                   round(loans, 1),
            "NIM (%)":                          round(nim * 100, 3),
            "Net Interest Income ($M)":         round(nii, 1),
            "Noninterest Income ($M)":          round(nonint_inc, 1),
            "Total Revenue ($M)":               round(total_rev, 1),
            "Efficiency Ratio (%)":             round(eff_ratio * 100, 2),
            "Noninterest Expense ($M)":         round(nonint_exp, 1),
            "Provision for Credit Losses ($M)": round(provision, 1),
            "Net Income ($M)":                  round(net_inc, 1),
            "FTE Count":                        fte,
            "Total Comp Expense ($M)":          round(total_comp, 1),
            "FTE per $B Loans":                 round(fte / (loans / 1_000), 2)
        })
    return pd.DataFrame(rows)

PRESETS = {
    "Rising Rate":  dict(nim_change_bps=15,  loan_growth_rate=0.060, gdp_growth=0.025,
                         unemployment_delta=-0.3, base_charge_off=0.0035),
    "Flat Rate":    dict(nim_change_bps=0,   loan_growth_rate=0.045, gdp_growth=0.022,
                         unemployment_delta=0.0,  base_charge_off=0.0045),
    "Falling Rate": dict(nim_change_bps=-20, loan_growth_rate=0.030, gdp_growth=0.005,
                         unemployment_delta=0.8,  base_charge_off=0.0065),
}
COLORS = {
    "Rising Rate":  "#2ecc71",
    "Flat Rate":    "#3498db",
    "Falling Rate": "#e74c3c",
    "Custom":       "#9b59b6"
}

with st.sidebar:
    st.title("Retail Banking Budget Simulator")
    st.markdown("---")
    scenario_choice = st.radio(
        "Select Macro Scenario",
        ["Rising Rate", "Flat Rate", "Falling Rate", "Custom"]
    )
    st.markdown("---")
    if scenario_choice == "Custom":
        st.subheader("Custom Assumptions")
        fed_funds   = st.slider("Fed Funds Rate (%)", 0.0, 6.0, 2.5, 0.25)
        loan_growth = st.slider("Loan Growth Rate (%)", -5.0, 10.0, 4.5, 0.5) / 100
        gdp_growth  = st.slider("GDP Growth Rate (%)", 0.0, 5.0, 2.2, 0.25) / 100
        nim_bps     = (fed_funds - 2.5) * 12
        active_params = dict(nim_change_bps=nim_bps, loan_growth_rate=loan_growth,
                             gdp_growth=gdp_growth, unemployment_delta=0.0,
                             base_charge_off=0.0045)
    else:
        active_params = PRESETS[scenario_choice]

    n_years = st.selectbox("Planning Horizon", [1, 2, 3], index=2)
    st.markdown("---")
    st.caption("Data: FRED (Federal Reserve), BLS NAICS 5221.")
    st.caption("Simulated division — not proprietary PNC data.")

df = project_scenario(**active_params, n_years=n_years)
df_display = df[df["Year"].isin([f"Year {i}" for i in range(1, n_years + 1)])]
yr_last = df_display.iloc[-1]

st.title("Retail Banking Budget Cycle Simulator")
st.markdown("**3-Year Strategic Plan | Macro Scenario Analysis | Staffing Forecast**")
st.markdown(f"**Active Scenario: {scenario_choice}**")
st.markdown("---")

k1, k2, k3, k4 = st.columns(4)
yr_n_ni  = yr_last["Net Income ($M)"]
rev_cagr = ((yr_last["Total Revenue ($M)"] / total_revenue_M) ** (1 / n_years) - 1) * 100
k1.metric(f"Year {n_years} Net Income",  f"${yr_n_ni:.1f}M",
          f"{((yr_n_ni / net_income_base_M) - 1) * 100:+.1f}% vs Y0")
k2.metric("Revenue CAGR",               f"{rev_cagr:.2f}%")
k3.metric(f"Year {n_years} Efficiency", f"{yr_last['Efficiency Ratio (%)']:.1f}%")
k4.metric(f"Year {n_years} FTE",        f"{yr_last['FTE Count']:,}",
          f"{yr_last['FTE Count'] - fte_base:+,} vs Y0")

st.markdown("---")
st.subheader("Net Income Projection")
fig1 = go.Figure()
fig1.add_trace(go.Bar(
    x=df_display["Year"],
    y=df_display["Net Income ($M)"],
    marker_color=COLORS.get(scenario_choice, "#9b59b6"),
    text=[f"${v:.1f}M" for v in df_display["Net Income ($M)"]],
    textposition="outside"
))
fig1.update_layout(
    template="plotly_white",
    height=380,
    yaxis_title="Net Income ($M)",
    xaxis_title="Year"
)
st.plotly_chart(fig1, use_container_width=True)

st.subheader("Staffing Forecast")
fig2 = go.Figure()
fte_vals  = [round(fte_base / (AVG_LOANS_M / 1_000), 2)] + df_display["FTE per $B Loans"].tolist()
yr_labels = ["Year 0"] + df_display["Year"].tolist()
fig2.add_trace(go.Scatter(
    x=yr_labels,
    y=fte_vals,
    mode="lines+markers+text",
    line=dict(color=COLORS.get(scenario_choice, "#9b59b6"), width=2.5),
    text=[f"{v:.1f}" for v in fte_vals],
    textposition="top center"
))
fig2.add_hrect(
    y0=18, y1=24,
    fillcolor="rgba(0,200,0,0.08)",
    line_width=0,
    annotation_text="Optimal Range: 18-24 FTE per B loans"
)
fig2.add_hline(y=18, line_dash="dash", line_color="gray", annotation_text="18 FTE floor")
fig2.add_hline(y=24, line_dash="dash", line_color="gray", annotation_text="24 FTE ceiling")
fig2.update_layout(
    template="plotly_white",
    height=380,
    yaxis_title="FTE per $B Loans",
    xaxis_title="Year"
)
st.plotly_chart(fig2, use_container_width=True)

st.subheader("Rate Sensitivity — Year 3 Net Income Impact")
base_ni    = project_scenario(**PRESETS["Flat Rate"], n_years=3).iloc[-1]["Net Income ($M)"]
shocks_bps = [-100, -50, 0, 50, 100]
heat_vals, heat_labels = [], []
for shock in shocks_bps:
    p = PRESETS["Flat Rate"].copy()
    p["nim_change_bps"] += shock * 0.12
    ni = project_scenario(**p, n_years=3).iloc[-1]["Net Income ($M)"]
    heat_vals.append(round((ni - base_ni) / base_ni * 100, 2))
    heat_labels.append(f"{shock:+d} bps")

fig3 = go.Figure(go.Heatmap(
    z=[heat_vals],
    x=heat_labels,
    y=["Net Income Delta Pct"],
    colorscale="RdYlGn",
    zmid=0,
    text=[[f"{v:+.1f}%" for v in heat_vals]],
    texttemplate="%{text}",
    colorbar=dict(title="% Change")
))
fig3.update_layout(template="plotly_white", height=220)
st.plotly_chart(fig3, use_container_width=True)

st.subheader("Management Commentary")
narratives = {
    "Rising Rate": (
        "Under the Rising Rate scenario the division benefits from asset-sensitive balance sheet "
        "positioning with NIM expanding approximately 15bps per year as loan yields reprice faster "
        "than deposit costs. Strong loan growth of 6% annually drives revenue CAGR above the "
        "flat-rate base case while lower unemployment supports below-cycle charge-off rates. "
        "Staffing demand grows proportionally with volume keeping FTE productivity within the "
        "18 to 24 FTE per billion dollar benchmark band."
    ),
    "Flat Rate": (
        "The Flat Rate scenario represents the base case — a stable mid-cycle environment with "
        "NIM holding near 3.1%, moderate loan growth of 4.5%, and charge-offs at the historical "
        "mid-cycle level of 0.45%. Operating leverage drives a gradual improvement in the "
        "efficiency ratio as fixed costs grow more slowly than revenue. This scenario is the "
        "planning anchor for budget submission."
    ),
    "Falling Rate": (
        "The Falling Rate scenario models a stress environment driven by rate cuts in response "
        "to economic weakness. NIM compresses 20bps per year, loan demand weakens to 3% growth, "
        "and rising unemployment pushes charge-off rates above 0.65% by Year 3. Net income "
        "growth is materially constrained relative to the base case. Management should evaluate "
        "expense actions to partially offset revenue headwinds."
    ),
    "Custom": (
        "Custom scenario active. Key drivers reflect user-defined assumptions for the Fed Funds "
        "rate, loan growth, and GDP trajectory. Review KPI cards above for projected outcomes. "
        "Use the sidebar sliders to stress-test planning assumptions in real time."
    )
}
st.info(narratives.get(scenario_choice, narratives["Custom"]))

with st.expander("Full Projected Income Statement"):
    st.dataframe(df_display.set_index("Year"), use_container_width=True)

st.subheader("Download Plan")

def build_excel(df_out, scenario_name):
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = scenario_name[:31]
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")
    afill = PatternFill("solid", fgColor="D9E1F2")
    ws_out.cell(row=1, column=1,
                value=f"{scenario_name} — 3-Year Plan").font = Font(bold=True, size=13)
    for ci, col in enumerate(df_out.columns, 1):
        c = ws_out.cell(row=2, column=ci, value=col)
        c.font = hfont
        c.fill = hfill
    for ri, row_data in enumerate(
        dataframe_to_rows(df_out, index=False, header=False), 3
    ):
        for ci, val in enumerate(row_data, 1):
            c = ws_out.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0:
                c.fill = afill
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf.getvalue()

excel_bytes = build_excel(df_display, scenario_choice)
st.download_button(
    label=f"Download {scenario_choice} Plan (.xlsx)",
    data=excel_bytes,
    file_name=f"retail_banking_{scenario_choice.lower().replace(' ', '_')}_plan.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.markdown("---")
st.caption(
    "Retail Banking Budget Cycle Simulator | Python, Streamlit, Plotly | "
    "Data: FRED Federal Reserve, BLS NAICS 5221 | "
    "Simulated division — does not represent proprietary PNC Financial Services data."
)
